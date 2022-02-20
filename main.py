import datetime
import schedule
import sqlite3 as sql
import time
import winsound
import threading
from openpyxl import load_workbook

"""
TO DO:

LINK GUI TO BACKEND
LINK TREEVIEW TO DB
ADD ACTION BUTTONS
PLAY SIREN FUNCTION
PAY SOS SIREN FUNCTION

"""

# This Function deletes all entries on the DB and stops all scheduled alarms


def clean():
    schedule.clear()
    connect = sql.connect('events3.db')
    cursor = connect.cursor()
    cursor.execute(f"SELECT * FROM EVENTS")
    for element in cursor.fetchall():
        task = list(element)
        cursor.execute(f"""delete from EVENTS where ID='{task[0]}'""")
    connect.commit()
    connect.close()
    return 0


# This Function loads data from the local db


def load_db():
    monday = []
    tuesday = []
    wednesday = []
    thursday = []
    friday = []
    saturday = []
    sunday = []
    connect = sql.connect('events3.db')
    cursor = connect.cursor()
    cursor.execute(f"SELECT * FROM EVENTS")

    for element in cursor.fetchall():
        task = list(element)
        if task[1] == 'Monday':
            monday.append(task)
        elif task[1] == 'Tuesday':
            tuesday.append(task)
        elif task[1] == 'Wednesday':
            wednesday.append(task)
        elif task[1] == 'Thursday':
            thursday.append(task)
        elif task[1] == 'Friday':
            friday.append(task)
        elif task[1] == 'Saturday':
            saturday.append(task)
        elif task[1] == 'Sunday':
            sunday.append(task)
    connect.close()
    return monday, tuesday, wednesday, thursday, friday, saturday, sunday


# PLAYS SIREN


def alarm():
    print('ALARM!')


# This Function counts the number of alarms on each sheet of the imported xlsx
def count_alarms(alarm_column):
    number_of_alarms = 0
    for cell in alarm_column:
       if cell[0].value != None:
           number_of_alarms += 1
       else:
           return number_of_alarms+1


# This function loads Alarms from the xlsx file and adds them to DB and schedule
def load_alarms(workbook_path):
    clean()
    c = sql.connect('events3.db')
    cur = c.cursor()
    wb = load_workbook(workbook_path)
    sheets = []
    for sheet in wb.sheetnames:
        sheets.append(sheet)

    MondayAlarms = wb[sheets[0]]['A2':'A128']
    a = count_alarms(MondayAlarms)
    MondayAlarms = list(wb[sheets[0]]['A2':f'A{a}'])
    MondayDescription = list(wb[sheets[0]]['B2':f'B{a}'])

    for x in range(0, a-1):
        cur.execute(f"INSERT INTO EVENTS(DAY, START_TIME, DESCRIPTION) VALUES ('Monday','{MondayAlarms[x][0].value}', '{MondayDescription[x][0].value}')")
        lastadded = cur.execute(f"SELECT * FROM EVENTS where ID={cur.lastrowid}")
        lastadded = list(lastadded)
        lastadded = lastadded[0]
        schedule.every().monday.at(lastadded[2]).do(alarm).tag(f"t{lastadded[0]}")

    TuesdayAlarms = wb[sheets[1]]['A2':'A128']
    a = count_alarms(TuesdayAlarms)
    TuesdayAlarms = list(wb[sheets[1]]['A2':f'A{a}'])
    TuesdayDescription = list(wb[sheets[1]]['B2':f'B{a}'])

    for x in range(0, a-1):
        cur.execute(f"INSERT INTO EVENTS(DAY, START_TIME, DESCRIPTION) VALUES ('Tuesday','{TuesdayAlarms[x][0].value}', '{TuesdayDescription[x][0].value}')")
        lastadded = cur.execute(f"SELECT * FROM EVENTS where ID={cur.lastrowid}")
        lastadded = list(lastadded)
        lastadded = lastadded[0]
        schedule.every().tuesday.at(lastadded[2]).do(alarm).tag(f"t{lastadded[0]}")

    WednesdayAlarms = wb[sheets[2]]['A2':'A128']
    a = count_alarms(WednesdayAlarms)
    WednesdayAlarms = list(wb[sheets[2]]['A2':f'A{a}'])
    WednesdayDescription = list(wb[sheets[2]]['B2':f'B{a}'])

    for x in range(0, a-1):
        cur.execute(f"INSERT INTO EVENTS(DAY, START_TIME, DESCRIPTION) VALUES ('Wednesday','{WednesdayAlarms[x][0].value}', '{WednesdayDescription[x][0].value}')")
        lastadded = cur.execute(f"SELECT * FROM EVENTS where ID={cur.lastrowid}")
        lastadded = list(lastadded)
        lastadded = lastadded[0]
        schedule.every().wednesday.at(lastadded[2]).do(alarm).tag(f"t{lastadded[0]}")

    ThursdayAlarms = wb[sheets[3]]['A2':'A128']
    a = count_alarms(ThursdayAlarms)
    ThursdayAlarms = list(wb[sheets[3]]['A2':f'A{a}'])
    ThursdayDescription = list(wb[sheets[3]]['B2':f'B{a}'])

    for x in range(0, a-1):
        cur.execute(f"INSERT INTO EVENTS(DAY, START_TIME, DESCRIPTION) VALUES ('Thursday','{ThursdayAlarms[x][0].value}', '{ThursdayDescription[x][0].value}')")
        lastadded = cur.execute(f"SELECT * FROM EVENTS where ID={cur.lastrowid}")
        lastadded = list(lastadded)
        lastadded = lastadded[0]
        schedule.every().thursday.at(lastadded[2]).do(alarm).tag(f"t{lastadded[0]}")

    FridayAlarms = wb[sheets[4]]['A2':'A128']
    a = count_alarms(FridayAlarms)
    FridayAlarms = list(wb[sheets[4]]['A2':f'A{a}'])
    FridayDescription = list(wb[sheets[4]]['B2':f'B{a}'])

    for x in range(0, a-1):
        cur.execute(f"INSERT INTO EVENTS(DAY, START_TIME, DESCRIPTION) VALUES ('Friday','{FridayAlarms[x][0].value}', '{FridayDescription[x][0].value}')")
        lastadded = cur.execute(f"SELECT * FROM EVENTS where ID={cur.lastrowid}")
        lastadded = list(lastadded)
        lastadded = lastadded[0]
        schedule.every().friday.at(lastadded[2]).do(alarm).tag(f"t{lastadded[0]}")

    SaturdayAlarms = wb[sheets[5]]['A2':'A128']
    a = count_alarms(SaturdayAlarms)
    SaturdayAlarms = list(wb[sheets[5]]['A2':f'A{a}'])
    SaturdayDescription = list(wb[sheets[5]]['B2':f'B{a}'])

    for x in range(0, a-1):
        cur.execute(f"INSERT INTO EVENTS(DAY, START_TIME, DESCRIPTION) VALUES ('Saturday','{SaturdayAlarms[x][0].value}', '{SaturdayDescription[x][0].value}')")
        lastadded = cur.execute(f"SELECT * FROM EVENTS where ID={cur.lastrowid}")
        lastadded = list(lastadded)
        lastadded = lastadded[0]
        schedule.every().saturday.at(lastadded[2]).do(alarm).tag(f"t{lastadded[0]}")

    SundayAlarms = wb[sheets[6]]['A2':'A128']
    a = count_alarms(SundayAlarms)
    SundayAlarms = list(wb[sheets[6]]['A2':f'A{a}'])
    SundayDescription = list(wb[sheets[6]]['B2':f'B{a}'])

    for x in range(0, a-1):
        cur.execute(f"INSERT INTO EVENTS(DAY, START_TIME, DESCRIPTION) VALUES ('Sunday','{SundayAlarms[x][0].value}', '{SundayDescription[x][0].value}')")
        lastadded = cur.execute(f"SELECT * FROM EVENTS where ID={cur.lastrowid}")
        lastadded = list(lastadded)
        lastadded = lastadded[0]
        schedule.every().sunday.at(lastadded[2]).do(alarm).tag(f"t{lastadded[0]}")
        #print(f"{SundayAlarms[x][0].value}:{SundayDescription[x][0].value}")

    c.commit()
    c.close()
    wb.close()

    return 0


# get current day name
def today():
    now = datetime.datetime.now()
    return now.strftime("%A")


# this function removes the alarm from db and schedule
def kill(id):
    job = schedule.get_jobs(f"t{id}")
    schedule.cancel_job(job[0])
    c = sql.connect('events3.db')
    cur = c.cursor()
    cur.execute(f"""delete from EVENTS where ID='{id}'""")
    c.commit()
    c.close()
    return 0


# this function adds the alarm to db and schedule
def add(day, time, description='PAS DE DESCRIPTION'):
    c = sql.connect('events3.db')
    cur = c.cursor()
    cur.execute(f"INSERT INTO EVENTS(DAY, START_TIME, DESCRIPTION) VALUES ('{day}','{time}', '{description}')")
    lastadded = cur.execute(f"SELECT * FROM EVENTS where ID={cur.lastrowid}")

    lastadded = list(lastadded)
    lastadded = lastadded[0]
    if lastadded[1] == 'Monday':
        schedule.every().monday.at(lastadded[2]).do(alarm).tag(f"t{lastadded[0]}")
    elif lastadded[1] == 'Tuesday':
        schedule.every().tuesday.at(lastadded[2]).do(alarm).tag(f"t{lastadded[0]}")
    elif lastadded[1] == 'Wednesday':
        schedule.every().wednesday.at(lastadded[2]).do(alarm).tag(f"t{lastadded[0]}")
    elif lastadded[1] == 'Thursday':
        schedule.every().thursday.at(lastadded[2]).do(alarm).tag(f"t{lastadded[0]}")
    elif lastadded[1] == 'Friday':
        schedule.every().friday.at(lastadded[2]).do(alarm).tag(f"t{lastadded[0]}")
    elif lastadded[1] == 'Saturday':
        schedule.every().saturday.at(lastadded[2]).do(alarm).tag(f"t{lastadded[0]}")
    elif lastadded[1] == 'Sunday':
        schedule.every().sunday.at(lastadded[2]).do(alarm).tag(f"t{lastadded[0]}")

    c.commit()
    c.close()
    return 0


# This function loads Alarms from the DB and adds them to schedule at startup
def initialize():
    schedule.clear()
    monday, tuesday, wednesday, thursday, friday, saturday, sunday = load_db()

    for day in monday:
        schedule.every().monday.at(day[2]).do(alarm).tag(f"t{day[0]}")
        print(f"t{day[0]}")
    for day in tuesday:
        schedule.every().tuesday.at(day[2]).do(alarm).tag(f"t{day[0]}")
    for day in wednesday:
        schedule.every().wednesday.at(day[2]).do(alarm).tag(f"t{day[0]}")
    for day in thursday:
        schedule.every().thursday.at(day[2]).do(alarm).tag(f"t{day[0]}")
    for day in friday:
        schedule.every().friday.at(day[2]).do(alarm).tag(f"t{day[0]}")
    for day in saturday:
        schedule.every().saturday.at(day[2]).do(alarm).tag(f"t{day[0]}")
    for day in sunday:
        schedule.every().sunday.at(day[2]).do(alarm).tag(f"t{day[0]}")

    all_jobs = schedule.get_jobs()
    print(all_jobs)

    while True:
        schedule.run_pending()
        time.sleep(1)


# Main Function to be removed later
def main():
    add('Sunday', '02:16')
    threading.Thread(target=initialize).start()
    all_jobs = schedule.get_jobs()
    print(len(all_jobs))
    # add('Saturday', '20:11', 'buh')
    all_jobs = schedule.get_jobs()
    print(len(all_jobs))


main()

"""
all_jobs = schedule.get_jobs()
print(all_jobs)
kill(35)
"""

# add("Thursday", "03:54", "TESTING")

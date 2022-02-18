from openpyxl import load_workbook
import sqlite3 as sql
import schedule


def alarm():
    print('ALARM!')


def count_alarms(alarm_column):
    number_of_alarms = 0
    for cell in alarm_column:
       if cell[0].value != None:
           number_of_alarms += 1
       else:
           return number_of_alarms+1


def load_alarms(workbook_path):
    c = sql.connect('events3.db')
    cur = c.cursor()
    wb = load_workbook(workbook_path)
    sheets = []
    for sheet in wb.sheetnames:
        sheets.append(sheet)

    MondayAlarms = wb[sheets[0]]['A2':'A128']
    a = count_alarms(MondayAlarms)
    MondayAlarms = list(wb[sheets[0]]['A2':f'A{a}'])
    MondayDescription = list(wb[sheets[0]]['B2':f'B{a}'])

    for x in range(0, a-1):
        cur.execute(f"INSERT INTO EVENTS(DAY, START_TIME, DESCRIPTION) VALUES ('Monday','{MondayAlarms[x][0].value}', '{MondayDescription[x][0].value}')")
        lastadded = cur.execute(f"SELECT * FROM EVENTS where ID={cur.lastrowid}")
        lastadded = list(lastadded)
        lastadded = lastadded[0]
        schedule.every().monday.at(lastadded[2]).do(alarm).tag(f"t{lastadded[0]}")

    TuesdayAlarms = wb[sheets[1]]['A2':'A128']
    a = count_alarms(TuesdayAlarms)
    TuesdayAlarms = list(wb[sheets[1]]['A2':f'A{a}'])
    TuesdayDescription = list(wb[sheets[1]]['B2':f'B{a}'])

    for x in range(0, a-1):
        cur.execute(f"INSERT INTO EVENTS(DAY, START_TIME, DESCRIPTION) VALUES ('Tuesday','{TuesdayAlarms[x][0].value}', '{TuesdayDescription[x][0].value}')")
        lastadded = cur.execute(f"SELECT * FROM EVENTS where ID={cur.lastrowid}")
        lastadded = list(lastadded)
        lastadded = lastadded[0]
        schedule.every().tuesday.at(lastadded[2]).do(alarm).tag(f"t{lastadded[0]}")

    WednesdayAlarms = wb[sheets[2]]['A2':'A128']
    a = count_alarms(WednesdayAlarms)
    WednesdayAlarms = list(wb[sheets[2]]['A2':f'A{a}'])
    WednesdayDescription = list(wb[sheets[2]]['B2':f'B{a}'])

    for x in range(0, a-1):
        cur.execute(f"INSERT INTO EVENTS(DAY, START_TIME, DESCRIPTION) VALUES ('Wednesday','{WednesdayAlarms[x][0].value}', '{WednesdayDescription[x][0].value}')")
        lastadded = cur.execute(f"SELECT * FROM EVENTS where ID={cur.lastrowid}")
        lastadded = list(lastadded)
        lastadded = lastadded[0]
        schedule.every().wednesday.at(lastadded[2]).do(alarm).tag(f"t{lastadded[0]}")

    ThursdayAlarms = wb[sheets[3]]['A2':'A128']
    a = count_alarms(ThursdayAlarms)
    ThursdayAlarms = list(wb[sheets[3]]['A2':f'A{a}'])
    ThursdayDescription = list(wb[sheets[3]]['B2':f'B{a}'])

    for x in range(0, a-1):
        cur.execute(f"INSERT INTO EVENTS(DAY, START_TIME, DESCRIPTION) VALUES ('Thursday','{ThursdayAlarms[x][0].value}', '{ThursdayDescription[x][0].value}')")
        lastadded = cur.execute(f"SELECT * FROM EVENTS where ID={cur.lastrowid}")
        lastadded = list(lastadded)
        lastadded = lastadded[0]
        schedule.every().thursday.at(lastadded[2]).do(alarm).tag(f"t{lastadded[0]}")

    FridayAlarms = wb[sheets[4]]['A2':'A128']
    a = count_alarms(FridayAlarms)
    FridayAlarms = list(wb[sheets[4]]['A2':f'A{a}'])
    FridayDescription = list(wb[sheets[4]]['B2':f'B{a}'])

    for x in range(0, a-1):
        cur.execute(f"INSERT INTO EVENTS(DAY, START_TIME, DESCRIPTION) VALUES ('Friday','{FridayAlarms[x][0].value}', '{FridayDescription[x][0].value}')")
        lastadded = cur.execute(f"SELECT * FROM EVENTS where ID={cur.lastrowid}")
        lastadded = list(lastadded)
        lastadded = lastadded[0]
        schedule.every().friday.at(lastadded[2]).do(alarm).tag(f"t{lastadded[0]}")

    SaturdayAlarms = wb[sheets[5]]['A2':'A128']
    a = count_alarms(SaturdayAlarms)
    SaturdayAlarms = list(wb[sheets[5]]['A2':f'A{a}'])
    SaturdayDescription = list(wb[sheets[5]]['B2':f'B{a}'])

    for x in range(0, a-1):
        cur.execute(f"INSERT INTO EVENTS(DAY, START_TIME, DESCRIPTION) VALUES ('Saturday','{SaturdayAlarms[x][0].value}', '{SaturdayDescription[x][0].value}')")
        lastadded = cur.execute(f"SELECT * FROM EVENTS where ID={cur.lastrowid}")
        lastadded = list(lastadded)
        lastadded = lastadded[0]
        schedule.every().saturday.at(lastadded[2]).do(alarm).tag(f"t{lastadded[0]}")

    SundayAlarms = wb[sheets[6]]['A2':'A128']
    a = count_alarms(SundayAlarms)
    SundayAlarms = list(wb[sheets[6]]['A2':f'A{a}'])
    SundayDescription = list(wb[sheets[6]]['B2':f'B{a}'])

    for x in range(0, a-1):
        cur.execute(f"INSERT INTO EVENTS(DAY, START_TIME, DESCRIPTION) VALUES ('Sunday','{SundayAlarms[x][0].value}', '{SundayDescription[x][0].value}')")
        lastadded = cur.execute(f"SELECT * FROM EVENTS where ID={cur.lastrowid}")
        lastadded = list(lastadded)
        lastadded = lastadded[0]
        schedule.every().sunday.at(lastadded[2]).do(alarm).tag(f"t{lastadded[0]}")
        #print(f"{SundayAlarms[x][0].value}:{SundayDescription[x][0].value}")

    c.commit()
    c.close()
    wb.close()

    return 0

load_alarms("import.xlsx")
#print(count_alarms(Sunday))
